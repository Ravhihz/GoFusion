from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from .models import Dataset, Tweet, PreprocessedTweet
import json


def preliminary_check_view(request, dataset_id):
    """
    Tampilkan tweets untuk dicek relevansinya (manual check)
    User mark each tweet as relevant/not relevant
    """
    try:
        dataset = get_object_or_404(Dataset, id=dataset_id)
        
        # Get statistics
        total_tweets = Tweet.objects.filter(dataset_id=dataset_id).count()
        checked_tweets = Tweet.objects.filter(dataset_id=dataset_id, relevance_checked=True).count()
        relevant_tweets = Tweet.objects.filter(dataset_id=dataset_id, is_relevant=True, relevance_checked=True).count()
        not_relevant_tweets = Tweet.objects.filter(dataset_id=dataset_id, is_relevant=False, relevance_checked=True).count()
        
        progress = (checked_tweets / total_tweets * 100) if total_tweets > 0 else 0
        
        # Get unchecked tweets (10 per page for better UX)
        page = int(request.GET.get('page', 1))
        per_page = 50
        offset = (page - 1) * per_page
        
        unchecked_tweets = Tweet.objects.filter(
            dataset_id=dataset_id,
            relevance_checked=False
        ).select_related('preprocessed').order_by('id')[offset:offset + per_page]
        
        # Check if there are more pages
        total_unchecked = Tweet.objects.filter(dataset_id=dataset_id, relevance_checked=False).count()
        has_next_page = total_unchecked > (page * per_page)
        
        # ✅ FIXED CONTEXT - Match template variable names!
        context = {
            'dataset': dataset,
            'unchecked_tweets': unchecked_tweets,      # ✅ Match template
            'total_tweets': total_tweets,              # ✅ Match template
            'checked_count': checked_tweets,           # ✅ Match template
            'relevant_count': relevant_tweets,         # ✅ Match template
            'not_relevant_count': not_relevant_tweets, # ✅ Match template
            'progress': round(progress, 2),
            'page': page,
            'has_next_page': has_next_page,
            'total_unchecked': total_unchecked,
        }
        
        return render(request, 'core/preliminary_check.html', context)
        
    except Exception as e:
        messages.error(request, f"Error loading preliminary check: {str(e)}")
        return redirect('core:dataset_list')


def mark_relevance_ajax(request):
    """
    API endpoint untuk mark tweet sebagai relevant/not relevant
    Called via AJAX from frontend
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            tweet_id = data.get('tweet_id')
            is_relevant = data.get('is_relevant')
            note = data.get('note', '')
            
            if not tweet_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Tweet ID is required'
                }, status=400)
            
            # Get tweet
            tweet = Tweet.objects.get(id=tweet_id)
            
            # Update relevance
            tweet.is_relevant = is_relevant
            tweet.relevance_checked = True
            tweet.relevance_note = note
            tweet.checked_by = request.user.username if request.user.is_authenticated else 'manual'
            tweet.checked_at = timezone.now()
            tweet.save()
            
            # Update dataset statistics
            dataset = tweet.dataset
            dataset.update_statistics()
            
            return JsonResponse({
                'success': True,
                'message': f"Tweet marked as {'relevant' if is_relevant else 'not relevant'}",
                'tweet_id': tweet_id
            })
            
        except Tweet.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Tweet not found'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    }, status=405)


def bulk_mark_relevance(request, dataset_id):
    """
    Bulk mark multiple tweets at once
    Untuk mempercepat proses preliminary check
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            tweet_ids = data.get('tweet_ids', [])
            is_relevant = data.get('is_relevant')
            
            if not tweet_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'No tweet IDs provided'
                }, status=400)
            
            # Update all tweets
            updated = Tweet.objects.filter(
                id__in=tweet_ids,
                dataset_id=dataset_id
            ).update(
                is_relevant=is_relevant,
                relevance_checked=True,
                checked_by=request.user.username if request.user.is_authenticated else 'bulk',
                checked_at=timezone.now()
            )
            
            # Update dataset statistics
            dataset = Dataset.objects.get(id=dataset_id)
            dataset.update_statistics()
            
            return JsonResponse({
                'success': True,
                'message': f"Successfully marked {updated} tweets",
                'updated_count': updated
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    }, status=405)


def preliminary_statistics(request, dataset_id):
    """
    Get detailed statistics for preliminary check
    """
    try:
        dataset = get_object_or_404(Dataset, id=dataset_id)
        
        total = Tweet.objects.filter(dataset_id=dataset_id).count()
        checked = Tweet.objects.filter(dataset_id=dataset_id, relevance_checked=True).count()
        unchecked = total - checked
        
        relevant = Tweet.objects.filter(dataset_id=dataset_id, is_relevant=True, relevance_checked=True).count()
        not_relevant = Tweet.objects.filter(dataset_id=dataset_id, is_relevant=False, relevance_checked=True).count()
        
        # Get preprocessed tweets for additional stats
        preprocessed_count = PreprocessedTweet.objects.filter(tweet__dataset_id=dataset_id).count()
        
        stats = {
            'dataset_name': dataset.name,
            'total_tweets': total,
            'checked_tweets': checked,
            'unchecked_tweets': unchecked,
            'relevant_tweets': relevant,
            'not_relevant_tweets': not_relevant,
            'preprocessed_tweets': preprocessed_count,
            'progress_percentage': round((checked / total * 100), 2) if total > 0 else 0,
            'relevance_rate': round((relevant / checked * 100), 2) if checked > 0 else 0,
        }
        
        return JsonResponse({
            'success': True,
            'statistics': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def skip_preliminary_check(request, dataset_id):
    """
    Skip preliminary check - mark all tweets as relevant
    Use case: kalau user yakin semua data sudah bersih
    """
    if request.method == 'POST':
        try:
            dataset = get_object_or_404(Dataset, id=dataset_id)
            
            # Mark all unchecked tweets as relevant
            updated = Tweet.objects.filter(
                dataset_id=dataset_id,
                relevance_checked=False
            ).update(
                is_relevant=True,
                relevance_checked=True,
                checked_by='auto',
                checked_at=timezone.now(),
                relevance_note='Auto-marked as relevant (skipped manual check)'
            )
            
            # Update dataset statistics
            dataset.update_statistics()
            
            messages.success(request, f"Marked all {updated} unchecked tweets as relevant")
            
            return JsonResponse({
                'success': True,
                'message': f"Successfully marked {updated} tweets as relevant",
                'updated_count': updated
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    }, status=405)


@require_http_methods(["POST"])
def reset_preliminary_check(request, dataset_id):
    """
    Reset preliminary check results
    
    This will:
    - Clear keyword filter results (is_relevant, relevance_checked)
    - Reset tweets marked by 'keyword_filter'
    - Keep manual check results untouched
    - Allow user to re-apply filter with different keywords
    """
    try:
        dataset = get_object_or_404(Dataset, id=dataset_id)
        
        # Count tweets before reset
        before_count = Tweet.objects.filter(
            dataset=dataset,
            checked_by='keyword_filter'
        ).count()
        
        if before_count == 0:
            return JsonResponse({
                'success': False,
                'message': 'No preliminary check results to reset. Apply keyword filter first.'
            }, status=400)
        
        # Reset only tweets marked by keyword_filter
        # DON'T touch manually checked tweets
        reset_count = Tweet.objects.filter(
            dataset=dataset,
            checked_by='keyword_filter'
        ).update(
            is_relevant=True,  # Reset to default
            relevance_checked=False,  # Mark as unchecked
            checked_by='',
            checked_at=None,
            relevance_note=''
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Preliminary check reset successfully. {reset_count} tweets ready for re-filtering.',
            'reset_count': reset_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to reset: {str(e)}'
        }, status=500)


def export_preliminary_results(request, dataset_id):
    """
    Export preliminary check results to Excel
    """
    try:
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill
        
        dataset = get_object_or_404(Dataset, id=dataset_id)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Preliminary Check Results' # type: ignore
        
        # Header
        headers = ['Tweet ID', 'Username', 'Text', 'Is Relevant', 'Note', 'Checked By', 'Checked At']
        ws.append(headers) # type: ignore
        
        # Style header
        for cell in ws[1]: # type: ignore
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Get checked tweets
        tweets = Tweet.objects.filter(
            dataset_id=dataset_id,
            relevance_checked=True
        ).order_by('checked_at')
        
        # Add data
        for tweet in tweets:
            ws.append([ # type: ignore
                tweet.tweet_id,
                tweet.username,
                tweet.text[:100],
                'Yes' if tweet.is_relevant else 'No',
                tweet.relevance_note,
                tweet.checked_by,
                tweet.checked_at.strftime('%Y-%m-%d %H:%M:%S') if tweet.checked_at else ''
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15 # type: ignore
        ws.column_dimensions['B'].width = 15 # type: ignore
        ws.column_dimensions['C'].width = 50 # type: ignore
        ws.column_dimensions['D'].width = 12 # type: ignore
        ws.column_dimensions['E'].width = 30 # type: ignore
        ws.column_dimensions['F'].width = 15 # type: ignore
        ws.column_dimensions['G'].width = 20 # type: ignore
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=preliminary_check_{dataset.name}.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Error exporting results: {str(e)}")
        return redirect('preliminary_check', dataset_id=dataset_id)
    
# ADD THESE FUNCTIONS TO views_preliminary.py

def preview_keyword_filter(request, dataset_id):
    """Preview how many tweets will be affected by keyword filter"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            keywords = data.get('keywords', [])
            
            if not keywords:
                return JsonResponse({
                    'success': False,
                    'message': 'No keywords provided'
                }, status=400)
            
            dataset = get_object_or_404(Dataset, id=dataset_id)
            
            # Build query for tweets containing ANY keyword
            query = Q()
            for keyword in keywords:
                if keyword.strip():
                    query |= Q(text__icontains=keyword.strip())
            
            # Count unchecked tweets with keywords (will be relevant)
            relevant_count = Tweet.objects.filter(
                dataset_id=dataset_id,
                relevance_checked=False
            ).filter(query).count()
            
            # Count unchecked tweets without keywords (will be not relevant)
            total_unchecked = Tweet.objects.filter(
                dataset_id=dataset_id,
                relevance_checked=False
            ).count()
            
            not_relevant_count = total_unchecked - relevant_count
            
            # Count already checked
            unchanged_count = Tweet.objects.filter(
                dataset_id=dataset_id,
                relevance_checked=True
            ).count()
            
            return JsonResponse({
                'success': True,
                'relevant_count': relevant_count,
                'not_relevant_count': not_relevant_count,
                'unchanged_count': unchanged_count,
                'total_affected': relevant_count + not_relevant_count,
                'keywords_used': keywords
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    }, status=405)


def apply_keyword_filter(request, dataset_id):
    """
    Apply keyword filter - mark tweets as relevant/not relevant based on keywords
    
    Logic:
    - Tweets WITH keywords → is_relevant=True (will go to Manual Check)
    - Tweets WITHOUT keywords → is_relevant=False (excluded from workflow)
    - relevance_checked stays FALSE (Manual Check will set it to True later)
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        keywords = data.get('keywords', [])
        mark_rest = data.get('mark_rest', True)
        
        if not keywords:
            return JsonResponse({
                'success': False,
                'message': 'No keywords provided'
            }, status=400)
        
        dataset = get_object_or_404(Dataset, id=dataset_id)
        
        # Clean keywords
        keywords = [k.strip() for k in keywords if k.strip()]
        
        # Build query for tweets containing ANY keyword (case-insensitive)
        query = Q()
        for keyword in keywords:
            query |= Q(text__icontains=keyword)
        
        # Get tweets that haven't been checked yet
        unchecked_tweets = Tweet.objects.filter(
            dataset=dataset,
            relevance_checked=False  # Only process unchecked tweets
        )
        
        # Mark tweets WITH keywords as RELEVANT
        # NOTE: We only set is_relevant=True, NOT relevance_checked=True
        # Manual checking will set relevance_checked=True later
        relevant_updated = unchecked_tweets.filter(query).update(
            is_relevant=True,
            # relevance_checked stays False! ← PENTING!
            # These tweets will appear in Manual Check as "pending"
        )
        
        # Mark tweets WITHOUT keywords as NOT RELEVANT (if mark_rest=True)
        not_relevant_updated = 0
        if mark_rest:
            not_relevant_updated = unchecked_tweets.exclude(query).update(
                is_relevant=False,
                relevance_checked=True,  # ← Set True karena kita yakin ini NOT relevant
                checked_by='keyword_filter',
                checked_at=timezone.now(),
                relevance_note='Auto-excluded: no matching keywords'
            )
        
        return JsonResponse({
            'success': True,
            'message': f'Filter applied: {relevant_updated} relevant, {not_relevant_updated} excluded',
            'relevant_count': relevant_updated,
            'not_relevant_count': not_relevant_updated,
            'keywords_used': keywords
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
    
def preliminary_manual_check_view(request, dataset_id):
    try:
        dataset = get_object_or_404(Dataset, id=dataset_id)

        total_manual = Tweet.objects.filter(
            dataset_id=dataset_id,
            is_relevant=True
        ).count()

        checked_manual = Tweet.objects.filter(
            dataset_id=dataset_id,
            is_relevant=True,
            relevance_checked=True
        ).count()

        relevant_tweets = Tweet.objects.filter(
            dataset_id=dataset_id,
            is_relevant=True,
            relevance_checked=True
        ).count()

        not_relevant_tweets = Tweet.objects.filter(
            dataset_id=dataset_id,
            is_relevant=False,
            relevance_checked=True
        ).count()

        progress = (checked_manual / total_manual * 100) if total_manual > 0 else 0

        page = int(request.GET.get('page', 1))
        per_page = 50
        offset = (page - 1) * per_page

        unchecked_tweets = Tweet.objects.filter(
            dataset_id=dataset_id,
            is_relevant=True,
            relevance_checked=False
        ).select_related('preprocessed').order_by('id')[offset:offset + per_page]

        total_unchecked = Tweet.objects.filter(
            dataset_id=dataset_id,
            is_relevant=True,
            relevance_checked=False
        ).count()

        has_next_page = total_unchecked > (page * per_page)

        context = {
            'dataset': dataset,
            'unchecked_tweets': unchecked_tweets,
            'total_tweets': total_manual,
            'checked_count': checked_manual,
            'relevant_count': relevant_tweets,
            'not_relevant_count': not_relevant_tweets,
            'progress': round(progress, 2),
            'page': page,
            'has_next_page': has_next_page,
            'total_unchecked': total_unchecked,
        }

        return render(request, 'core/preliminary_manual.html', context)

    except Exception as e:
        messages.error(request, f"Error loading manual review: {str(e)}")
        return redirect('core:dataset_list')
